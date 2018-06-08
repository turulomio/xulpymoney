#!/usr/bin/python3
## @package libxlsxgenerator
## @brief Este módulo permite la lectura y escritura de ficheros xlsx de Microsoft Excel
## En el caso de que sea necesaria la modificación de este módulo deberá realizarse en el proyecto Xulpymoney y ser luego copiado al proyecto en el que se necesite

import datetime
import openpyxl
import openpyxl.comments
import openpyxl.cell
import openpyxl.styles
import openpyxl.worksheet
import openpyxl.formatting.rule
import os
import platform

from libodfgenerator import columnAdd, makedirs, rowAdd

class OpenPyXL:
    def __init__(self,filename,template=None):
        if template==None:
            if platform.system()=="Windows":
                self.filename=filename.replace("/home/informatica/", "z:/")
            else:
                self.filename=filename
            self.wb=openpyxl.Workbook()
        else:
            if platform.system()=="Windows":
                self.filename=filename.replace("/home/informatica/", "z:/")
                self.template=template.replace("/usr/share/pysgae/", "Templates/")
            else:
                self.filename=filename
                self.template=template
            self.wb=openpyxl.load_workbook(self.template)
        
        self.ws_current_id=id
        
        self.stNormal=openpyxl.styles.Style( 
            alignment=openpyxl.styles.Alignment(horizontal='right', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=False), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin') 
            ), 
            number_format="#,##0"
        )        
        self.stNormalDecimal=openpyxl.styles.Style( 
            alignment=openpyxl.styles.Alignment(horizontal='right', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=False), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin') 
            ), 
            number_format="#,##0.00"
        )
        
        self.stNormalEuros=openpyxl.styles.Style( 
            alignment=openpyxl.styles.Alignment(horizontal='right', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=False), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin') 
            ), 
            number_format="#,##0.00 €"
        )        
        
        self.stNormalLeft=openpyxl.styles.Style( 
            alignment=openpyxl.styles.Alignment(horizontal='left', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=False), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin') 
            ), 
            number_format="#,##0"
        )
        
        self.stNormalCenter=openpyxl.styles.Style( 
            alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=False), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin') 
            ),
        )
        self.stOrange=openpyxl.styles.Style(
            fill=openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('FFFFDCA8')), 
            alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=True), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin') 
            )
        )
        self.stYellow=openpyxl.styles.Style(
            fill=openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('FFFFFFC0')), 
            alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=True), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin') 
            )
        )
        self.stGreen=openpyxl.styles.Style(
            fill=openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('FFC0FFC0')), 
            alignment=openpyxl.styles.Alignment(horizontal='left', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=True), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin') 
            )
        )
        self.stGrey1=openpyxl.styles.Style(
            fill=openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('FFDCDCDC')), 
            alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=True), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin') 
            )
        )
        self.stGrey2=openpyxl.styles.Style(
            fill=openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('FFC3C3C3')), 
            alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=True), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin') 
            )
        )
        
        self.stGrey1Number=openpyxl.styles.Style(
            fill=openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('FFDCDCDC')), 
            alignment=openpyxl.styles.Alignment(horizontal='right', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=True), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin')
            ), 
            number_format="#,##0"
        )
        self.stGrey1Euros=openpyxl.styles.Style(
            fill=openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('FFDCDCDC')), 
            alignment=openpyxl.styles.Alignment(horizontal='right', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=True), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin')
            ), 
            number_format="#,##0.00 €"
        )        
    
        self.stGrey2Number=openpyxl.styles.Style(
            fill=openpyxl.styles.PatternFill(patternType='solid', fgColor=openpyxl.styles.Color('FFC3C3C3')), 
            alignment=openpyxl.styles.Alignment(horizontal='right', vertical='center'), 
            font=openpyxl.styles.Font(name='Arial', size=10, bold=True), 
            border=openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin'),
                top=openpyxl.styles.Side(border_style='thin'),
                right=openpyxl.styles.Side(border_style='thin'),
                bottom=openpyxl.styles.Side(border_style='thin') 
            )
        )
    def freezePanels(self, cell):
        """Cell is string coordinates"""
        ws=self.get_sheet_by_id(self.ws_current_id)
        ws.freeze_panes=ws.cell(cell)
        
    def setSelectedCell(self, cell):
        """Cell is string coordinates
        
        Estaq función fue echa a modo prueba error"""
        ws=self.get_sheet_by_id(self.ws_current_id)
        ws.sheet_view.pane.topLeftCell=cell
        ws.sheet_view.selection=[]
        ws.sheet_view.selection.append(openpyxl.worksheet.Selection("topLeft", cell, None, cell))
        

    def setSheetName(self, name):
        """Changes current id"""
        ws=self.get_sheet_by_id(self.ws_current_id)
        ws.title=name

    def createSheet(self, name):
        """Create a sheet at the end, renames it and selects it as current"""
        self.wb.create_sheet(title=name)
        self.ws_current_id=self.get_sheet_id(name)
        
    def setColorScale(self, range):
        ws=self.get_sheet_by_id(self.ws_current_id)
        ws.conditional_formatting.add(range, 
                            openpyxl.formatting.rule.ColorScaleRule(
                                                start_type='percentile', start_value=0, start_color='00FF00',
                                                mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                                                end_type='percentile', end_value=100, end_color='FF0000'
                                                )   
                                            )

    def sheet_name(self, id=None):
        if id==None:
            id=self.ws_current_id
        return self.wb.get_sheet_names()[id]

    def get_sheet_by_id(self, id):
        return self.wb.get_sheet_by_name(self.wb.get_sheet_names()[id])
        
    def get_sheet_id(self, name):
        for id, s_name in enumerate(self.wb.get_sheet_names()):
            if s_name==name:
                return id
        return None
        
    def remove_sheet_by_id(self, id):
        ws=self.get_sheet_by_id(id)
        self.wb.remove_sheet(ws)
        self.get_sheet_by_id(0)

    def save(self, filename=None):
        if filename==None:
            filename=self.filename
        makedirs(os.path.dirname(filename))
        self.wb.save(filename)
        
        if os.path.exists(filename)==False:
            print( "*** ERROR: El fichero no se ha generado ***")
        
        

    def overwrite(self, letter,number, result, ws=None, style=None):
        """Writes in aself.ws_current worksheet from cell, (letter, number) the array result"""
        def setStyle(cell, value, style):
            if style==None:#Auto
                if value.__class__ in(datetime.datetime, str):
                    cell.style=self.stNormalLeft
                else:
                    cell.style=self.stNormal
            else:
                cell.style=style
        def setValue(cell, value, style):        
            if value.__class__ in (str, int, float):#Un solo valor
                cell.value=value
                setStyle(cell, value, style)
            elif value.__class__ in (datetime.datetime, ):
                cell.value=value
                setStyle(cell, value, style)
                cell.number_format="YYYY-MM-DD HH:mm"
            elif value.__class__ in (datetime.date, ):
                cell.value=value
                setStyle(cell, value, style)
                cell.number_format="YYYY-MM-DD"
            elif value==None:
                pass
            else:
                print(value.__class__, "VALUE CLASS NOT FOUND")
        ################################################
        if ws==None:
            ws=self.get_sheet_by_id(self.ws_current_id)
            
        if result.__class__ in (str, int, float, datetime.datetime):#Un solo valor
            c=ws[letter+number]
            setValue(c, result, style)
        elif result.__class__ in (list,):#Una lista
            for i,row in enumerate(result):
                if row.__class__ in (int, str, float, datetime.datetime):#Una lista de una columna
                    c=ws[letter+rowAdd(number,i)]
                    setValue(c, result[i], style)
                elif row.__class__ in (list, ):#Una lista de varias columnas
                    for j,column in enumerate(row):
                        c=ws[columnAdd(letter, j)+rowAdd(number,i)]
                        setValue(c, result[i][j], style)
                else:
                    print(row.__class__, "ROW CLASS NOT FOUND",  row)

    def setCellName(self, range, name):
        """Cell name to use in formulas"""
        ws=self.get_sheet_by_id(self.ws_current_id)
        self.wb.create_named_range(name, ws, range)

    def setCurrentSheet(self, id):
        self.ws_current_id=id
        
    def setColumnsWidth(self, arrWidths):
        """arrWidths es un array de anchuras"""
        ws=self.get_sheet_by_id(self.ws_current_id)
        for i in range(len(arrWidths)):
            ws.column_dimensions[openpyxl.cell.get_column_letter(i+1)].width=arrWidths[i]
   
    def mergeCells(self, range, style=None):
        """range='A1:B1
        
        style was added to avoid formating errors after merging
        
        """
        ws=self.get_sheet_by_id(self.ws_current_id)
        if style!=None:
            for row in ws.iter_rows(range):
                for c in row:
                    c.style=style
        ws.merge_cells(range)
        
        
    def setComment(self, cell, comment):
        """Cell is string coordinates"""
        ws=self.get_sheet_by_id(self.ws_current_id)
        ws[cell].comment=openpyxl.comments.Comment(comment, "PySGAE")
